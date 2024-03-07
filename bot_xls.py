
import telebot
from openpyxl import load_workbook
bot = telebot.TeleBot("6849059285:AAFhQAZugM5rzqrmPeiCctBZYQ-0VbVFsWg")





TEXT = ['ДА']

key = []
key1 = 1











import random
import time
import timeit
import xlsxwriter
from openpyxl import load_workbook
wb = load_workbook('./rand.xlsx')
sheet = wb.get_sheet_by_name('УСПЕХ')
rand = {}
i = 1
sheet.title
sheet['A1'].value
c = sheet['B2']
c.row
c.column
c.coordinate
sheet.cell(row=1, column=2).value
a = 2
key = 1
while (sheet.cell(row=i, column=2).value) != None:
    i += 1
    for_in_motive = i
for i in range(a, i):
    key = a
    rand[key] = (sheet.cell(row=a, column=2).value)
    a += 1


# print(i)
#
# for key, value in rand.items():
#     print(key, '\t - \t', value, sep='')
motive = ''

def import_v_rand():
    key1 = random.randint(1, for_in_motive)
    motive = (rand[key1])
    print(f'Доброго времени суток \n  {motive} \nВсего самого наилучшего')






@bot.message_handler(commands=["start"])
def generate_keys(message):
    import_v_rand()
    key1 = random.randint(1, for_in_motive)
    motive = (rand[key1])
    bot.send_message(message.chat.id, f'\n  {motive} \n\nВсего самого наилучшего')


bot.infinity_polling()



















